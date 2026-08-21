"""
V2X, Inc. (NYSE: VVX) — Single-name sponsor LBO model.

Builds output/V2X_LBO.xlsx via openpyxl (all calc tabs use live formulas,
inputs live only on the Assumptions tab) and prints a returns/credit-stats
summary to the terminal using an independent Python replica of the same
formulas (openpyxl does not evaluate formulas, so this is how we verify the
workbook's logic and satisfy the "print to terminal" deliverable).

All historical figures are sourced to V2X 10-K filings; see data/*.csv for
the citation attached to every line item pulled from SEC EDGAR (CIK 1601548).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.properties import CalcProperties
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# 0. SINGLE SOURCE OF TRUTH: assumptions + historicals as Python constants.
#    These drive BOTH the Excel Assumptions tab and the Python numeric
#    replica used for the terminal printout, so the two never diverge.
# ---------------------------------------------------------------------------

# --- Historical actuals, $000s, sourced to 10-Ks (see data/*.csv) ---
HIST_YEARS = [2021, 2022, 2023, 2024, 2025]
REVENUE = {2021: 1783665, 2022: 2890860, 2023: 3963126, 2024: 4322155, 2025: 4480038}
COGS = {2021: 1623245, 2022: 2595848, 2023: 3628271, 2024: 3979193, 2025: 4106656}
SGA = {2021: 98400, 2022: 239241, 2023: 210439, 2024: 183758, 2025: 179112}
OP_INCOME = {2021: 62020, 2022: 55771, 2023: 124416, 2024: 159204, 2025: 194270}
LOSS_EXT_DEBT = {2021: 0, 2022: 0, 2023: -22298, 2024: -1998, 2025: -2527}
INTEREST_EXP = {2021: -7985, 2022: -61879, 2023: -122442, 2024: -107900, 2025: -79909}
OTHER_EXP = {2021: 0, 2022: 0, 2023: -4194, 2024: -10465, 2025: -10931}
PRETAX = {2021: 54035, 2022: -6108, 2023: -24518, 2024: 38841, 2025: 100903}
TAX_EXP = {2021: 8307, 2022: 8222, 2023: -1945, 2024: 4157, 2025: 23021}
NET_INCOME = {2021: 45728, 2022: -14330, 2023: -22573, 2024: 34684, 2025: 77882}

DEP_EXP = {2021: 6526, 2022: 13472, 2023: 22408, 2024: 20747, 2025: 16984}
AMORT_INTANG = {2021: 10028, 2022: 48643, 2023: 90423, 2024: 90821, 2025: 90621}
AMORT_CLOUD = {2021: 0, 2022: 0, 2023: 480, 2024: 3314, 2025: 4919}
CAPEX = {2021: 9776, 2022: 12425, 2023: 25021, 2024: 11787, 2025: 11923}
CFO = {2021: 61339, 2022: 93495, 2023: 187968, 2024: 254237, 2025: 181992}

BS_YEARS = [2022, 2023, 2024, 2025]
CASH = {2022: 116067, 2023: 72651, 2024: 268321, 2025: 368994}
RECEIVABLES = {2022: 728582, 2023: 705995, 2024: 710068, 2025: 738922}
PREPAID_OTHER_CA = {2022: 42309 + 44974, 2023: 49242 + 46981, 2024: 124081, 2025: 127102}
TOTAL_CA = {2022: 931932, 2023: 874869, 2024: 1102470, 2025: 1235018}
PPE_NET = {2022: 78715, 2023: 85429, 2024: 62001, 2025: 52383}
GOODWILL = {2022: 1653822, 2023: 1656926, 2024: 1656926, 2025: 1677154}
INTANGIBLES = {2022: 497951, 2023: 407530, 2024: 323068, 2025: 239760}
OTHER_NCA = {2022: 17858 + 52825, 2023: 15931 + 41215, 2024: 84378, 2025: 76525}
TOTAL_ASSETS = {2022: 3233103, 2023: 3081900, 2024: 3228843, 2025: 3280840}
AP = {2022: 406706, 2023: 453052, 2024: 547568, 2025: 557042}
COMP_BENEFITS = {2022: 168038, 2023: 158088, 2024: 166918, 2025: 176530}
ST_DEBT = {2022: 11850, 2023: 15361, 2024: 20003, 2025: 14935}
OTHER_ACCRUED = {2022: 196538, 2023: 213700, 2024: 261735, 2025: 267039}
TOTAL_CL = {2022: 783132, 2023: 840201, 2024: 996224, 2025: 1015546}
LT_DEBT = {2022: 1262811, 2023: 1100269, 2024: 1087484, 2025: 1083234}
DEF_TAX_LIAB = {2022: 15813, 2023: 11763, 2024: 20983, 2025: 28357}
OTHER_NCL = {2022: 133185 + 41083, 2023: 104176 + 34691, 2024: 98000, 2025: 69067}
TOTAL_LIAB = {2022: 2236024, 2023: 2091100, 2024: 2202691, 2025: 2196204}
TOTAL_EQUITY = {2022: 997079, 2023: 990800, 2024: 1026152, 2025: 1084636}

# --- Debt at close (12/31/2025), $000s, sourced to 10-K Note DEBT ---
EXISTING_DEBT_PRINCIPAL = 1123819  # per maturity schedule, R69
EXISTING_CASH = 368994

# --- Transaction assumptions ---
ENTRY_MULTIPLE = 11.0
EXIT_MULTIPLE = 11.0
TXN_FEE_PCT = 0.015
FINANCING_FEE_PCT = 0.025
MIN_CASH = 50000.0
HOLD_YEARS = 5
PROJ_YEARS = [2026, 2027, 2028, 2029, 2030]

# --- Financing structure at close ---
TLB_TURNS = 4.50
NOTES_TURNS = 1.50
REVOLVER_CAPACITY = 150000.0
TLB_RATE = 0.0700
NOTES_RATE = 0.0850
REVOLVER_RATE = 0.0725
REVOLVER_COMMIT_FEE = 0.00375
TLB_MAND_AMORT_PCT = 0.01  # % of original principal, per annum
CASH_SWEEP_PCT = 0.75

# --- Operating assumptions ---
REV_GROWTH = [0.040, 0.040, 0.045, 0.045, 0.045]
EBITDA_MARGIN = [0.070, 0.072, 0.074, 0.075, 0.076]
DA_PCT_REV = 0.024
CAPEX_PCT_REV = 0.0035
NWC_PCT_REV = -0.030
TAX_RATE = 0.25

# --- Market reference (informational only) ---
MKT_PRICE = 79.01
MKT_SHARES = 31350.0
MKT_DATE = "2026-08-20"

ENTRY_EBITDA = OP_INCOME[2025] + DEP_EXP[2025] + AMORT_INTANG[2025] + AMORT_CLOUD[2025]  # 306,794

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
BLUE = Font(name="Calibri", size=11, color="FF1F4E96")
BLACK = Font(name="Calibri", size=11, color="FF000000")
GREEN = Font(name="Calibri", size=11, color="FF1E7B34")
BOLD_BLACK = Font(name="Calibri", size=11, color="FF000000", bold=True)
BOLD_WHITE = Font(name="Calibri", size=12, color="FFFFFFFF", bold=True)
ITALIC_GRAY = Font(name="Calibri", size=10, color="FF7F7F7F", italic=True)
TITLE_FILL = PatternFill("solid", fgColor="FF1F4E96")
SECTION_FILL = PatternFill("solid", fgColor="FFDCE6F1")
CHECK_OK_FILL = PatternFill("solid", fgColor="FFC6EFCE")
CHECK_BAD_FILL = PatternFill("solid", fgColor="FFFFC7CE")
THIN = Side(style="thin", color="FFB7B7B7")
BORDER_BOTTOM = Border(bottom=THIN)
BORDER_TOP = Border(top=THIN)
BORDER_BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_MM = '#,##0.0;(#,##0.0)'
FMT_MM0 = '#,##0;(#,##0)'
FMT_USD_MM = '$#,##0.0;($#,##0.0)'
FMT_PCT = '0.0%'
FMT_PCT2 = '0.00%'
FMT_X = '0.00"x"'
FMT_USD_SH = '$#,##0.00'
FMT_YEAR = '0'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_bar(ws, text, span=8, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD_WHITE
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = TITLE_FILL


def subtitle(ws, text, row, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = ITALIC_GRAY


def section_header(ws, text, row, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color="FF1F4E96")
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = BORDER_BOTTOM


def label(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD_BLACK if bold else BLACK
    c.alignment = Alignment(indent=indent)
    return c


def val(ws, row, col, value, font=BLACK, num_fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=font.name, size=font.size, color=font.color, bold=bold or font.bold)
    if num_fmt:
        c.number_format = num_fmt
    return c


def yr_headers(ws, row, years, start_col=2, extra_label=None):
    if extra_label:
        c = ws.cell(row=row, column=start_col - 1, value=extra_label)
        c.font = BOLD_BLACK
    for i, y in enumerate(years):
        c = ws.cell(row=row, column=start_col + i, value=f"FY{y}")
        c.font = BOLD_BLACK
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER_BOTTOM


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
# Interest is computed on beginning-of-year balances throughout the Debt Schedule
# (documented convention -- see tab subtitle), which avoids circularity entirely,
# so no iterative-calculation setting is required for this workbook to open and
# calculate correctly in Excel with default settings.

ws_a = wb.active
ws_a.title = "Assumptions"
ws_h = wb.create_sheet("Historical Financials")
ws_su = wb.create_sheet("Sources & Uses")
ws_d = wb.create_sheet("Debt Schedule")
ws_om = wb.create_sheet("Operating Model")
ws_fcf = wb.create_sheet("Free Cash Flow")
ws_r = wb.create_sheet("Returns")
ws_s = wb.create_sheet("Sensitivities")
ws_cs = wb.create_sheet("Credit Stats")

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

AR = {}  # Assumptions row map

# =============================================================================
# 1. ASSUMPTIONS TAB
# =============================================================================
set_col_widths(ws_a, [42, 13, 13, 13, 13, 13, 55])
title_bar(ws_a, "V2X, Inc. (NYSE: VVX) — LBO Model: Assumptions", span=7)
subtitle(ws_a, "All inputs in blue. $ in 000s unless noted. Change any blue cell to flex the model.", 2, span=7)

r = 4
section_header(ws_a, "Transaction Assumptions", r, span=7); r += 1
AR['entry_date'] = r; label(ws_a, r, "Entry / close date"); val(ws_a, r, 2, "12/31/2025", BLUE); r += 1
AR['entry_mult'] = r; label(ws_a, r, "Entry TEV / EBITDA multiple")
val(ws_a, r, 2, ENTRY_MULTIPLE, BLUE, FMT_X)
val(ws_a, r, 7, "11.0x = KBR/LinQuest precedent (11.0x); modest premium to VVX undisturbed trading multiple (~10.5x, see Market Reference below); within Perspecta/Veritas comp range of 10.5x-13.3x (Perspecta DEFM14A)", ITALIC_GRAY)
r += 1
AR['txn_fee_pct'] = r; label(ws_a, r, "Transaction fees (% of entry TEV)")
val(ws_a, r, 2, TXN_FEE_PCT, BLUE, FMT_PCT); r += 1
AR['fin_fee_pct'] = r; label(ws_a, r, "Financing fees (% of new debt raised)")
val(ws_a, r, 2, FINANCING_FEE_PCT, BLUE, FMT_PCT); r += 1
AR['min_cash'] = r; label(ws_a, r, "Minimum cash balance funded at close")
val(ws_a, r, 2, MIN_CASH, BLUE, FMT_MM0); r += 1
AR['hold_years'] = r; label(ws_a, r, "Hold period (years)")
val(ws_a, r, 2, HOLD_YEARS, BLUE, FMT_YEAR); r += 1
AR['exit_year'] = r; label(ws_a, r, "Exit year")
val(ws_a, r, 2, PROJ_YEARS[-1], BLUE, FMT_YEAR); r += 1
AR['exit_mult'] = r; label(ws_a, r, "Exit TEV / EBITDA multiple")
val(ws_a, r, 2, EXIT_MULTIPLE, BLUE, FMT_X)
val(ws_a, r, 7, "Base case assumes flat multiple (no re-rating); return relies on EBITDA growth + deleveraging. Flexed in Sensitivities.", ITALIC_GRAY)
r += 2

section_header(ws_a, "Financing Structure at Close", r, span=7); r += 1
AR['tlb_turns'] = r; label(ws_a, r, "New Term Loan B (x Entry EBITDA)")
val(ws_a, r, 2, TLB_TURNS, BLUE, FMT_X); r += 1
AR['notes_turns'] = r; label(ws_a, r, "New Senior Notes (x Entry EBITDA)")
val(ws_a, r, 2, NOTES_TURNS, BLUE, FMT_X); r += 1
AR['revolver_cap'] = r; label(ws_a, r, "Revolver capacity")
val(ws_a, r, 2, REVOLVER_CAPACITY, BLUE, FMT_MM0); r += 1
AR['tlb_rate'] = r; label(ws_a, r, "Term Loan B interest rate (all-in)")
val(ws_a, r, 2, TLB_RATE, BLUE, FMT_PCT2)
val(ws_a, r, 7, "vs. 6.50% effective rate on V2X's actual First Lien TLB at 12/31/25 (10-K Note DEBT); wider spread assumed for higher LBO leverage", ITALIC_GRAY)
r += 1
AR['notes_rate'] = r; label(ws_a, r, "Senior Notes interest rate (fixed)")
val(ws_a, r, 2, NOTES_RATE, BLUE, FMT_PCT2); r += 1
AR['rev_rate'] = r; label(ws_a, r, "Revolver interest rate (all-in, if drawn)")
val(ws_a, r, 2, REVOLVER_RATE, BLUE, FMT_PCT2); r += 1
AR['rev_commit_fee'] = r; label(ws_a, r, "Revolver commitment fee (undrawn)")
val(ws_a, r, 2, REVOLVER_COMMIT_FEE, BLUE, FMT_PCT2)
val(ws_a, r, 7, "V2X's actual 2025 Revolver unused fee is 0.25%-0.375% depending on leverage (10-K Note DEBT)", ITALIC_GRAY)
r += 1
AR['tlb_amort_pct'] = r; label(ws_a, r, "Term Loan B mandatory amortization (% of orig. principal, p.a.)")
val(ws_a, r, 2, TLB_MAND_AMORT_PCT, BLUE, FMT_PCT); r += 1
AR['sweep_pct'] = r; label(ws_a, r, "Cash flow sweep % (of FCF after mandatory amort.)")
val(ws_a, r, 2, CASH_SWEEP_PCT, BLUE, FMT_PCT); r += 2

section_header(ws_a, "Operating Assumptions — Revenue Growth & EBITDA Margin Path", r, span=7); r += 1
AR['path_header'] = r
yr_headers(ws_a, r, PROJ_YEARS, start_col=2, extra_label=None)
r += 1
AR['rev_growth_row'] = r; label(ws_a, r, "Revenue growth %")
for i, g in enumerate(REV_GROWTH):
    val(ws_a, r, 2 + i, g, BLUE, FMT_PCT)
r += 1
AR['ebitda_margin_row'] = r; label(ws_a, r, "EBITDA margin %")
for i, m in enumerate(EBITDA_MARGIN):
    val(ws_a, r, 2 + i, m, BLUE, FMT_PCT)
r += 2

section_header(ws_a, "Operating Assumptions — Flat Rates", r, span=7); r += 1
AR['da_pct'] = r; label(ws_a, r, "D&A (% of revenue)")
val(ws_a, r, 2, DA_PCT_REV, BLUE, FMT_PCT)
val(ws_a, r, 7, "vs. FY25A actual D&A/revenue of 2.51% (10-K cash flow statement)", ITALIC_GRAY)
r += 1
AR['capex_pct'] = r; label(ws_a, r, "Capex (% of revenue)")
val(ws_a, r, 2, CAPEX_PCT_REV, BLUE, FMT_PCT)
val(ws_a, r, 7, "vs. FY23A-FY25A actual capex/revenue of 0.27%-0.63% (10-K cash flow statements)", ITALIC_GRAY)
r += 1
AR['nwc_pct'] = r; label(ws_a, r, "Net working capital (% of revenue)")
val(ws_a, r, 2, NWC_PCT_REV, BLUE, FMT_PCT)
val(ws_a, r, 7, "Negative NWC is a source of cash as revenue grows; vs. FY25A actual of -3.00% of revenue (see Historical Financials tab)", ITALIC_GRAY)
r += 1
AR['tax_rate'] = r; label(ws_a, r, "Cash tax rate")
val(ws_a, r, 2, TAX_RATE, BLUE, FMT_PCT)
val(ws_a, r, 7, "vs. FY25A effective rate of 22.8% (10-K income statement); rounded up as a normalized go-forward estimate", ITALIC_GRAY)
r += 2

AR['mkt_ref_start'] = r  # Market Reference section body is inserted after Historical Financials is built (needs HR row refs)
r += 7
AR['assump_end'] = r

print("Assumptions tab (core) built. Continuing with Historical Financials...")

# =============================================================================
# 2. HISTORICAL FINANCIALS TAB
# =============================================================================
set_col_widths(ws_h, [42, 12, 12, 12, 12, 12, 3, 65])
title_bar(ws_h, "V2X, Inc. (NYSE: VVX) — Historical Financials", span=8)
subtitle(ws_h, "$ in 000s unless noted. Every figure traces to a 10-K per the Source column.", 2, span=8)

r = 4
label(ws_h, r, "Filing references used below:", bold=True); r += 1
label(ws_h, r, "  10-K FY2025 (period ended 12/31/2025, filed 2026-02-23), SEC accession 0001601548-26-000015", indent=1); r += 1
label(ws_h, r, "  10-K FY2023 (period ended 12/31/2023, filed 2024-03-05), SEC accession 0001601548-24-000004", indent=1); r += 1
r += 1

HR = {}

section_header(ws_h, "Consolidated Statements of Income (Loss) ($000s)", r, span=8); r += 1
yr_headers(ws_h, r, HIST_YEARS, start_col=2)
ws_h.cell(row=r, column=8, value="Source").font = BOLD_BLACK
r += 1

def hist_row(ws, r, label_text, data_dict, years, source_text, num_fmt=FMT_MM0, start_col=2, indent=0):
    label(ws, r, label_text, indent=indent)
    for i, y in enumerate(years):
        if y in data_dict:
            val(ws, r, start_col + i, data_dict[y], BLACK, num_fmt)
    ws.cell(row=r, column=8, value=source_text).font = ITALIC_GRAY
    ws.cell(row=r, column=8).alignment = Alignment(wrap_text=False)

SRC_IS = "10-K FY2025 Consolidated Statements of Income (Loss), FY23-25 cols; 10-K FY2023 Consolidated Statements of (Loss) Income, FY21-23 cols"
HR['revenue'] = r; hist_row(ws_h, r, "Revenue", REVENUE, HIST_YEARS, SRC_IS); r += 1
HR['cogs'] = r; hist_row(ws_h, r, "Cost of revenue", COGS, HIST_YEARS, SRC_IS); r += 1
HR['sga'] = r; hist_row(ws_h, r, "Selling, general & administrative", SGA, HIST_YEARS, SRC_IS); r += 1
HR['op_income'] = r; hist_row(ws_h, r, "Operating income", OP_INCOME, HIST_YEARS, SRC_IS); r += 1
HR['loss_ext'] = r; hist_row(ws_h, r, "Loss on extinguishment of debt", LOSS_EXT_DEBT, HIST_YEARS, SRC_IS); r += 1
HR['interest_exp'] = r; hist_row(ws_h, r, "Interest expense, net", INTEREST_EXP, HIST_YEARS, SRC_IS); r += 1
HR['other_exp'] = r; hist_row(ws_h, r, "Other expense, net", OTHER_EXP, HIST_YEARS, SRC_IS); r += 1
HR['pretax'] = r; hist_row(ws_h, r, "Pretax income (loss)", PRETAX, HIST_YEARS, SRC_IS); r += 1
HR['tax_exp'] = r; hist_row(ws_h, r, "Income tax expense (benefit)", TAX_EXP, HIST_YEARS, SRC_IS); r += 1
label(ws_h, r, "Net income (loss)", bold=True)
for i, y in enumerate(HIST_YEARS):
    val(ws_h, r, 2 + i, NET_INCOME[y], BOLD_BLACK, FMT_MM0)
ws_h.cell(row=r, column=8, value=SRC_IS).font = ITALIC_GRAY
HR['net_income'] = r; r += 2

section_header(ws_h, "Consolidated Balance Sheets ($000s)", r, span=8); r += 1
yr_headers(ws_h, r, BS_YEARS, start_col=3)
ws_h.cell(row=r, column=8, value="Source").font = BOLD_BLACK
r += 1
SRC_BS = "10-K FY2025 Consolidated Balance Sheets, FY24-25 cols; 10-K FY2023 Consolidated Balance Sheets, FY22-23 cols"
HR['cash'] = r; hist_row(ws_h, r, "Cash, cash equivalents & restricted cash", CASH, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['receivables'] = r; hist_row(ws_h, r, "Receivables", RECEIVABLES, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['prepaid_other_ca'] = r; hist_row(ws_h, r, "Prepaid & other current assets (incl. inventory FY22-23)", PREPAID_OTHER_CA, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['total_ca'] = r; hist_row(ws_h, r, "Total current assets", TOTAL_CA, BS_YEARS, SRC_BS, start_col=3, indent=1); r += 1
HR['ppe'] = r; hist_row(ws_h, r, "Property, plant & equipment, net", PPE_NET, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['goodwill'] = r; hist_row(ws_h, r, "Goodwill", GOODWILL, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['intangibles'] = r; hist_row(ws_h, r, "Intangible assets, net", INTANGIBLES, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['other_nca'] = r; hist_row(ws_h, r, "Other non-current assets (incl. ROU FY22-23)", OTHER_NCA, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['total_assets'] = r; hist_row(ws_h, r, "Total assets", TOTAL_ASSETS, BS_YEARS, SRC_BS, start_col=3, indent=1); r += 1
HR['ap'] = r; hist_row(ws_h, r, "Accounts payable", AP, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['comp_ben'] = r; hist_row(ws_h, r, "Compensation & other employee benefits", COMP_BENEFITS, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['st_debt'] = r; hist_row(ws_h, r, "Short-term debt", ST_DEBT, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['other_accrued'] = r; hist_row(ws_h, r, "Other accrued liabilities", OTHER_ACCRUED, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['total_cl'] = r; hist_row(ws_h, r, "Total current liabilities", TOTAL_CL, BS_YEARS, SRC_BS, start_col=3, indent=1); r += 1
HR['lt_debt'] = r; hist_row(ws_h, r, "Long-term debt, net", LT_DEBT, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['def_tax'] = r; hist_row(ws_h, r, "Deferred tax liabilities", DEF_TAX_LIAB, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['other_ncl'] = r; hist_row(ws_h, r, "Other non-current liabilities (incl. op. lease liab. FY22-23)", OTHER_NCL, BS_YEARS, SRC_BS, start_col=3); r += 1
HR['total_liab'] = r; hist_row(ws_h, r, "Total liabilities", TOTAL_LIAB, BS_YEARS, SRC_BS, start_col=3, indent=1); r += 1
HR['total_equity'] = r; hist_row(ws_h, r, "Total shareholders' equity", TOTAL_EQUITY, BS_YEARS, SRC_BS, start_col=3, indent=1); r += 1

HR['bs_check'] = r
label(ws_h, r, "Balance check (Assets − Liabilities − Equity, should = 0)", indent=1)
for i, y in enumerate(BS_YEARS):
    col = 3 + i
    cl = get_column_letter(col)
    f = f"={cl}{HR['total_assets']}-{cl}{HR['total_liab']}-{cl}{HR['total_equity']}"
    c = val(ws_h, r, col, f, BLACK, FMT_MM0)
    ws_h.conditional_formatting.add(
        f"{cl}{r}",
        CellIsRule(operator='notEqual', formula=['0'], fill=CHECK_BAD_FILL)
    )
    ws_h.conditional_formatting.add(
        f"{cl}{r}",
        CellIsRule(operator='equal', formula=['0'], fill=CHECK_OK_FILL)
    )
r += 2

section_header(ws_h, "Consolidated Statements of Cash Flows ($000s)", r, span=8); r += 1
yr_headers(ws_h, r, HIST_YEARS, start_col=2)
ws_h.cell(row=r, column=8, value="Source").font = BOLD_BLACK
r += 1
SRC_CF = "10-K FY2025 Consolidated Statements of Cash Flows, FY23-25 cols; 10-K FY2023 Consolidated Statements of Cash Flows, FY21-23 cols"
HR['dep_exp'] = r; hist_row(ws_h, r, "Depreciation expense", DEP_EXP, HIST_YEARS, SRC_CF); r += 1
HR['amort_intang'] = r; hist_row(ws_h, r, "Amortization of intangible assets", AMORT_INTANG, HIST_YEARS, SRC_CF); r += 1
HR['amort_cloud'] = r; hist_row(ws_h, r, "Amortization of cloud computing arrangements", AMORT_CLOUD, HIST_YEARS, SRC_CF); r += 1
HR['capex'] = r; hist_row(ws_h, r, "Capital expenditures (purchases of capital assets & intangibles)", CAPEX, HIST_YEARS, SRC_CF); r += 1
HR['cfo'] = r; hist_row(ws_h, r, "Net cash provided by operating activities", CFO, HIST_YEARS, SRC_CF); r += 2

section_header(ws_h, "Computed Metrics (formulas referencing rows above — not separately sourced)", r, span=8); r += 1
yr_headers(ws_h, r, HIST_YEARS, start_col=2)
r += 1
HR['ebitda'] = r
label(ws_h, r, "EBITDA (= Operating income + D&A)", bold=True)
for i, y in enumerate(HIST_YEARS):
    col = 2 + i
    cl = get_column_letter(col)
    f = f"={cl}{HR['op_income']}+{cl}{HR['dep_exp']}+{cl}{HR['amort_intang']}+{cl}{HR['amort_cloud']}"
    val(ws_h, r, col, f, BOLD_BLACK, FMT_MM0)
r += 1
HR['ebitda_margin'] = r
label(ws_h, r, "EBITDA margin %")
for i, y in enumerate(HIST_YEARS):
    col = 2 + i
    cl = get_column_letter(col)
    f = f"={cl}{HR['ebitda']}/{cl}{HR['revenue']}"
    val(ws_h, r, col, f, BLACK, FMT_PCT)
r += 1
HR['nwc'] = r
label(ws_h, r, "Net working capital (Receiv.+Prepaid&other−AP−Comp−Other accrued)")
for i, y in enumerate(BS_YEARS):
    col = 3 + i
    cl = get_column_letter(col)
    f = f"={cl}{HR['receivables']}+{cl}{HR['prepaid_other_ca']}-{cl}{HR['ap']}-{cl}{HR['comp_ben']}-{cl}{HR['other_accrued']}"
    val(ws_h, r, col, f, BLACK, FMT_MM0)
r += 1
HR['nwc_pct'] = r
label(ws_h, r, "Net working capital (% of revenue)")
for i, y in enumerate(BS_YEARS):
    col = 3 + i
    cl = get_column_letter(col)
    f = f"={cl}{HR['nwc']}/{cl}{HR['revenue']}"
    val(ws_h, r, col, f, BLACK, FMT_PCT)
r += 2

section_header(ws_h, "LBO Entry Reference — Funded Debt & Cash at Close (12/31/2025)", r, span=8); r += 1
HR['existing_debt'] = r
label(ws_h, r, "Existing funded debt principal (First Lien TLB + 2023/2025 Term Loan)")
val(ws_h, r, 2, EXISTING_DEBT_PRINCIPAL, BLACK, FMT_MM0)
ws_h.cell(row=r, column=8, value="10-K FY2025 Note DEBT — Schedule of Maturities of Term Facility (Details); ties to Short-term debt of $14,935K on balance sheet for the 2026 maturity").font = ITALIC_GRAY
r += 1
HR['existing_cash'] = r
label(ws_h, r, "Existing cash, cash equivalents & restricted cash")
f = f"=F{HR['cash']}"
val(ws_h, r, 2, f, GREEN, FMT_MM0)
r += 1
HR['existing_net_debt'] = r
label(ws_h, r, "Existing net debt (principal basis)", bold=True)
f = f"=B{HR['existing_debt']}-B{HR['existing_cash']}"
val(ws_h, r, 2, f, BOLD_BLACK, FMT_MM0)
r += 1
HR['ebitda_2025_ref'] = r
label(ws_h, r, "FY2025A EBITDA (entry EBITDA for the LBO)", bold=True)
f = f"=F{HR['ebitda']}"
val(ws_h, r, 2, f, GREEN, FMT_MM0, bold=True)
r += 1

print("Historical Financials tab built.")

# --- Back-fill Assumptions "Market Reference" section now that HR rows exist ---
r = AR['mkt_ref_start']
section_header(ws_a, "Market Reference (informational — not used in calc-tab formulas)", r, span=7); r += 1
AR['mkt_price'] = r; label(ws_a, r, "VVX stock price")
val(ws_a, r, 2, MKT_PRICE, BLUE, FMT_USD_SH)
val(ws_a, r, 7, f"stockanalysis.com quote, {MKT_DATE} close", ITALIC_GRAY)
r += 1
AR['mkt_shares'] = r; label(ws_a, r, "Shares outstanding (000s)")
val(ws_a, r, 2, MKT_SHARES, BLUE, FMT_MM0)
val(ws_a, r, 7, f"stockanalysis.com, {MKT_DATE}", ITALIC_GRAY)
r += 1
AR['mkt_cap'] = r; label(ws_a, r, "Implied current market capitalization")
val(ws_a, r, 2, f"=B{AR['mkt_price']}*B{AR['mkt_shares']}", BLACK, FMT_MM0)
r += 1
AR['mkt_net_debt'] = r; label(ws_a, r, "Existing net debt at 12/31/2025 (principal basis)")
val(ws_a, r, 2, f"='Historical Financials'!B{HR['existing_net_debt']}", GREEN, FMT_MM0)
r += 1
AR['mkt_tev'] = r; label(ws_a, r, "Implied current trading TEV")
val(ws_a, r, 2, f"=B{AR['mkt_cap']}+B{AR['mkt_net_debt']}", BLACK, FMT_MM0)
r += 1
AR['mkt_ev_ebitda'] = r; label(ws_a, r, "Implied current trading TEV / FY25A EBITDA")
val(ws_a, r, 2, f"=B{AR['mkt_tev']}/'Historical Financials'!B{HR['ebitda_2025_ref']}", BLACK, FMT_X)
r += 1

print("Assumptions Market Reference section back-filled.")

# =============================================================================
# 3. SOURCES & USES TAB  ($ in millions — all $000s inputs divided by 1000)
# =============================================================================
set_col_widths(ws_su, [46, 15, 3, 46, 15])
title_bar(ws_su, "V2X, Inc. (NYSE: VVX) — Sources & Uses", span=5)
subtitle(ws_su, "$ in millions. All cells are formulas referencing the Assumptions and Historical Financials tabs.", 2, span=5)

SU = {}
r = 4
section_header(ws_su, "Transaction Value", r, span=5); r += 1
SU['entry_ebitda'] = r; label(ws_su, r, "Entry EBITDA (FY2025A)")
val(ws_su, r, 2, f"='Historical Financials'!B{HR['ebitda_2025_ref']}/1000", GREEN, FMT_USD_MM); r += 1
SU['entry_mult'] = r; label(ws_su, r, "Entry TEV / EBITDA multiple")
val(ws_su, r, 2, f"=Assumptions!B{AR['entry_mult']}", GREEN, FMT_X); r += 1
SU['entry_tev'] = r; label(ws_su, r, "Entry TEV", bold=True)
val(ws_su, r, 2, f"=B{SU['entry_ebitda']}*B{SU['entry_mult']}", BOLD_BLACK, FMT_USD_MM); r += 1
SU['existing_net_debt'] = r; label(ws_su, r, "Less: existing net debt (principal basis)")
val(ws_su, r, 2, f"='Historical Financials'!B{HR['existing_net_debt']}/1000", GREEN, FMT_USD_MM); r += 1
SU['equity_purchase_price'] = r; label(ws_su, r, "Equity purchase price", bold=True)
val(ws_su, r, 2, f"=B{SU['entry_tev']}-B{SU['existing_net_debt']}", BOLD_BLACK, FMT_USD_MM); r += 2

section_header(ws_su, "Financing Structure (computed here, used throughout model)", r, span=5); r += 1
SU['new_tlb'] = r; label(ws_su, r, "New Term Loan B")
val(ws_su, r, 2, f"=B{SU['entry_ebitda']}*Assumptions!B{AR['tlb_turns']}", BLACK, FMT_USD_MM); r += 1
SU['new_notes'] = r; label(ws_su, r, "New Senior Notes")
val(ws_su, r, 2, f"=B{SU['entry_ebitda']}*Assumptions!B{AR['notes_turns']}", BLACK, FMT_USD_MM); r += 1
SU['new_debt_total'] = r; label(ws_su, r, "Total new debt raised", bold=True)
val(ws_su, r, 2, f"=B{SU['new_tlb']}+B{SU['new_notes']}", BOLD_BLACK, FMT_USD_MM); r += 1
SU['new_debt_leverage'] = r; label(ws_su, r, "Total new debt / Entry EBITDA")
val(ws_su, r, 2, f"=B{SU['new_debt_total']}/B{SU['entry_ebitda']}", BLACK, FMT_X); r += 2

section_header(ws_su, "Uses", r, span=5); r += 1
SU['use_equity'] = r; label(ws_su, r, "Equity purchase price")
val(ws_su, r, 2, f"=B{SU['equity_purchase_price']}", GREEN, FMT_USD_MM); r += 1
SU['use_refi'] = r; label(ws_su, r, "Refinance existing debt (gross principal)")
val(ws_su, r, 2, f"='Historical Financials'!B{HR['existing_debt']}/1000", GREEN, FMT_USD_MM); r += 1
SU['use_txn_fees'] = r; label(ws_su, r, "Transaction fees")
val(ws_su, r, 2, f"=B{SU['entry_tev']}*Assumptions!B{AR['txn_fee_pct']}", BLACK, FMT_USD_MM); r += 1
SU['use_fin_fees'] = r; label(ws_su, r, "Financing fees")
val(ws_su, r, 2, f"=B{SU['new_debt_total']}*Assumptions!B{AR['fin_fee_pct']}", BLACK, FMT_USD_MM); r += 1
SU['use_min_cash'] = r; label(ws_su, r, "Minimum cash funded at close")
val(ws_su, r, 2, f"=Assumptions!B{AR['min_cash']}/1000", GREEN, FMT_USD_MM); r += 1
SU['total_uses'] = r; label(ws_su, r, "Total Uses", bold=True)
val(ws_su, r, 2, f"=SUM(B{SU['use_equity']}:B{SU['use_min_cash']})", BOLD_BLACK, FMT_USD_MM)
for col in range(1, 3):
    ws_su.cell(row=r, column=col).border = BORDER_TOP
r += 2

section_header(ws_su, "Sources", r, span=5); r += 1
SU['src_cash'] = r; label(ws_su, r, "Existing cash swept from target balance sheet")
val(ws_su, r, 2, f"='Historical Financials'!B{HR['existing_cash']}/1000", GREEN, FMT_USD_MM); r += 1
SU['src_tlb'] = r; label(ws_su, r, "New Term Loan B")
val(ws_su, r, 2, f"=B{SU['new_tlb']}", BLACK, FMT_USD_MM); r += 1
SU['src_notes'] = r; label(ws_su, r, "New Senior Notes")
val(ws_su, r, 2, f"=B{SU['new_notes']}", BLACK, FMT_USD_MM); r += 1
SU['src_revolver'] = r; label(ws_su, r, "Revolver draw at close (undrawn)")
val(ws_su, r, 2, 0, BLACK, FMT_USD_MM); r += 1
SU['src_equity'] = r; label(ws_su, r, "Sponsor equity (plug)", bold=True)
val(ws_su, r, 2, f"=B{SU['total_uses']}-SUM(B{SU['src_cash']}:B{SU['src_revolver']})", BOLD_BLACK, FMT_USD_MM); r += 1
SU['total_sources'] = r; label(ws_su, r, "Total Sources", bold=True)
val(ws_su, r, 2, f"=SUM(B{SU['src_cash']}:B{SU['src_equity']})", BOLD_BLACK, FMT_USD_MM)
for col in range(1, 3):
    ws_su.cell(row=r, column=col).border = BORDER_TOP
r += 1
SU['check'] = r
label(ws_su, r, "CHECK: Total Sources − Total Uses (must = 0)", bold=True)
cf = f"=B{SU['total_sources']}-B{SU['total_uses']}"
val(ws_su, r, 2, cf, BOLD_BLACK, FMT_USD_MM)
ws_su.conditional_formatting.add(f"B{r}", CellIsRule(operator='notEqual', formula=['0'], fill=CHECK_BAD_FILL))
ws_su.conditional_formatting.add(f"B{r}", CellIsRule(operator='equal', formula=['0'], fill=CHECK_OK_FILL))
r += 2

section_header(ws_su, "Capitalization Summary", r, span=5); r += 1
SU['sponsor_equity_pct'] = r; label(ws_su, r, "Sponsor equity / Entry TEV")
val(ws_su, r, 2, f"=B{SU['src_equity']}/B{SU['entry_tev']}", BLACK, FMT_PCT); r += 1
SU['opening_net_debt'] = r; label(ws_su, r, "Opening net debt at close (new debt − minimum cash)", bold=True)
val(ws_su, r, 2, f"=B{SU['new_debt_total']}-B{SU['use_min_cash']}", BOLD_BLACK, FMT_USD_MM); r += 1
SU['opening_leverage'] = r; label(ws_su, r, "Opening total debt / Entry EBITDA")
val(ws_su, r, 2, f"=B{SU['new_debt_total']}/B{SU['entry_ebitda']}", BLACK, FMT_X); r += 1
SU['opening_net_leverage'] = r; label(ws_su, r, "Opening net debt / Entry EBITDA")
val(ws_su, r, 2, f"=B{SU['opening_net_debt']}/B{SU['entry_ebitda']}", BLACK, FMT_X); r += 1

print("Sources & Uses tab built.")

# =============================================================================
# 4/5/6. OPERATING MODEL, DEBT SCHEDULE, FREE CASH FLOW
#    These three tabs are circularly linked (interest <-> FCF <-> debt paydown),
#    resolved by the workbook's iterative-calculation setting (see CalcProperties
#    above). Row numbers are fixed constants below so all three tabs can
#    reference each other regardless of code write-order.
# =============================================================================
PROJ_COLS = ['C', 'D', 'E', 'F', 'G']   # FY2026..FY2030 on Operating Model / Debt Schedule / FCF
ASSUMP_COLS = ['B', 'C', 'D', 'E', 'F']  # FY2026..FY2030 on the Assumptions tab's path rows

# --- Fixed row map: Operating Model ---
OM = dict(header=5, revenue=6, rev_growth=7, ebitda_margin=8, ebitda=9, da=10, ebit=11,
          interest=12, ebt=13, taxes=14, ni=15, capex=18, nwc_bal=19, nwc_chg=20, nwc_cash=21)

# --- Fixed row map: Debt Schedule ---
DR = dict(header=5, beg_cash=8, cfads=9, mand_amort_tlb=10, mand_amort_notes=11, cf_before_revolver=12,
          rev_beg=15, rev_draw=16, rev_end=17, rev_avg=18, rev_int=19,
          tlb_beg=22, tlb_amort=23, tlb_sweep=24, tlb_end=25, tlb_avg=26, tlb_int=27,
          notes_beg=30, notes_amort=31, notes_end=32, notes_avg=33, notes_int=34,
          total_int=37, total_debt=38, end_cash=39, net_debt=40, check_close=42)

# --- Fixed row map: Free Cash Flow ---
FR = dict(header=5, ni=6, da=7, capex=8, nwc_cash=9, cfads=10, mand_amort=13, sweep=14, revolver=15,
          net_chg_cash=16, check=18)

HR_COL_2025_HIST = get_column_letter(2 + HIST_YEARS.index(2025))   # F
HR_COL_2025_BS = get_column_letter(3 + BS_YEARS.index(2025))       # F

# =============================================================================
# 4. OPERATING MODEL TAB
# =============================================================================
set_col_widths(ws_om, [40, 13, 13, 13, 13, 13, 13])
title_bar(ws_om, "V2X, Inc. (NYSE: VVX) — Operating Model", span=7)
subtitle(ws_om, "$ in millions. FY2025A shown for reference; FY2026-FY2030 are formulas off the Assumptions tab.", 2, span=7)

section_header(ws_om, "Projected P&L", OM['header'] - 1, span=7)
label(ws_om, OM['header'], "")
val(ws_om, OM['header'], 2, "FY2025A", BOLD_BLACK); ws_om.cell(row=OM['header'], column=2).alignment = Alignment(horizontal="center")
for i, y in enumerate(PROJ_YEARS):
    val(ws_om, OM['header'], 3 + i, f"FY{y}", BOLD_BLACK)
    ws_om.cell(row=OM['header'], column=3 + i).alignment = Alignment(horizontal="center")
    ws_om.cell(row=OM['header'], column=3 + i).border = BORDER_BOTTOM
ws_om.cell(row=OM['header'], column=2).border = BORDER_BOTTOM

label(ws_om, OM['revenue'], "Revenue")
val(ws_om, OM['revenue'], 2, f"='Historical Financials'!{HR_COL_2025_HIST}{HR['revenue']}/1000", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    prev_col = 'B' if i == 0 else PROJ_COLS[i - 1]
    val(ws_om, OM['revenue'], 3 + i, f"={prev_col}{OM['revenue']}*(1+Assumptions!{ASSUMP_COLS[i]}{AR['rev_growth_row']})", BLACK, FMT_USD_MM)

label(ws_om, OM['rev_growth'], "Revenue growth %")
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['rev_growth'], 3 + i, f"=Assumptions!{ASSUMP_COLS[i]}{AR['rev_growth_row']}", GREEN, FMT_PCT)

label(ws_om, OM['ebitda_margin'], "EBITDA margin %")
val(ws_om, OM['ebitda_margin'], 2, f"='Historical Financials'!{HR_COL_2025_HIST}{HR['ebitda_margin']}", GREEN, FMT_PCT)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['ebitda_margin'], 3 + i, f"=Assumptions!{ASSUMP_COLS[i]}{AR['ebitda_margin_row']}", GREEN, FMT_PCT)

label(ws_om, OM['ebitda'], "EBITDA", bold=True)
val(ws_om, OM['ebitda'], 2, f"='Historical Financials'!{HR_COL_2025_HIST}{HR['ebitda']}/1000", GREEN, FMT_USD_MM, bold=True)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['ebitda'], 3 + i, f"={col}{OM['revenue']}*{col}{OM['ebitda_margin']}", BOLD_BLACK, FMT_USD_MM)

label(ws_om, OM['da'], "D&A")
val(ws_om, OM['da'], 2,
    f"=('Historical Financials'!{HR_COL_2025_HIST}{HR['dep_exp']}+'Historical Financials'!{HR_COL_2025_HIST}{HR['amort_intang']}+'Historical Financials'!{HR_COL_2025_HIST}{HR['amort_cloud']})/1000",
    GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['da'], 3 + i, f"={col}{OM['revenue']}*Assumptions!$B${AR['da_pct']}", BLACK, FMT_USD_MM)

label(ws_om, OM['ebit'], "EBIT")
val(ws_om, OM['ebit'], 2, f"=B{OM['ebitda']}-B{OM['da']}", BLACK, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['ebit'], 3 + i, f"={col}{OM['ebitda']}-{col}{OM['da']}", BLACK, FMT_USD_MM)

label(ws_om, OM['interest'], "Interest expense (net)")
val(ws_om, OM['interest'], 2, f"='Historical Financials'!{HR_COL_2025_HIST}{HR['interest_exp']}/1000", GREEN, FMT_USD_MM)
val(ws_om, OM['interest'], 7, "FY2025A = actual pre-LBO capital structure (for reference only). FY26-30 link to new LBO capital structure on Debt Schedule.", ITALIC_GRAY)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['interest'], 3 + i, f"=-'Debt Schedule'!{col}{DR['total_int']}", GREEN, FMT_USD_MM)

label(ws_om, OM['ebt'], "EBT")
val(ws_om, OM['ebt'], 2, f"=B{OM['ebit']}+B{OM['interest']}", BLACK, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['ebt'], 3 + i, f"={col}{OM['ebit']}+{col}{OM['interest']}", BLACK, FMT_USD_MM)

label(ws_om, OM['taxes'], "Taxes")
val(ws_om, OM['taxes'], 2, f"=MAX(0,B{OM['ebt']})*Assumptions!$B${AR['tax_rate']}", BLACK, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['taxes'], 3 + i, f"=MAX(0,{col}{OM['ebt']})*Assumptions!$B${AR['tax_rate']}", BLACK, FMT_USD_MM)

label(ws_om, OM['ni'], "Net income", bold=True)
val(ws_om, OM['ni'], 2, f"=B{OM['ebt']}-B{OM['taxes']}", BOLD_BLACK, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['ni'], 3 + i, f"={col}{OM['ebt']}-{col}{OM['taxes']}", BOLD_BLACK, FMT_USD_MM)

section_header(ws_om, "Capex & Working Capital", OM['capex'] - 1, span=7)
label(ws_om, OM['capex'], "Capex")
val(ws_om, OM['capex'], 2, f"='Historical Financials'!{HR_COL_2025_HIST}{HR['capex']}/1000", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['capex'], 3 + i, f"={col}{OM['revenue']}*Assumptions!$B${AR['capex_pct']}", BLACK, FMT_USD_MM)

label(ws_om, OM['nwc_bal'], "Net working capital balance")
val(ws_om, OM['nwc_bal'], 2, f"='Historical Financials'!{HR_COL_2025_BS}{HR['nwc']}/1000", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['nwc_bal'], 3 + i, f"={col}{OM['revenue']}*Assumptions!$B${AR['nwc_pct']}", BLACK, FMT_USD_MM)

label(ws_om, OM['nwc_chg'], "Change in NWC balance")
for i, col in enumerate(PROJ_COLS):
    prev_col = 'B' if i == 0 else PROJ_COLS[i - 1]
    val(ws_om, OM['nwc_chg'], 3 + i, f"={col}{OM['nwc_bal']}-{prev_col}{OM['nwc_bal']}", BLACK, FMT_USD_MM)

label(ws_om, OM['nwc_cash'], "Cash impact from NWC (increase in NWC = cash use)")
for i, col in enumerate(PROJ_COLS):
    val(ws_om, OM['nwc_cash'], 3 + i, f"=-{col}{OM['nwc_chg']}", BLACK, FMT_USD_MM)

print("Operating Model tab built.")

# =============================================================================
# 5. DEBT SCHEDULE TAB
# =============================================================================
set_col_widths(ws_d, [46, 13, 13, 13, 13, 13, 13])
title_bar(ws_d, "V2X, Inc. (NYSE: VVX) — Debt Schedule", span=7)
subtitle(ws_d, "$ in millions. Interest computed on beginning-of-year balances (documented convention; avoids circularity -- no iterative calculation needed). Average balances shown as a memo only.", 2, span=7)

label(ws_d, DR['header'], "")
val(ws_d, DR['header'], 2, "At Close", BOLD_BLACK); ws_d.cell(row=DR['header'], column=2).alignment = Alignment(horizontal="center")
for i, y in enumerate(PROJ_YEARS):
    val(ws_d, DR['header'], 3 + i, f"FY{y}", BOLD_BLACK)
    ws_d.cell(row=DR['header'], column=3 + i).alignment = Alignment(horizontal="center")
    ws_d.cell(row=DR['header'], column=3 + i).border = BORDER_BOTTOM
ws_d.cell(row=DR['header'], column=2).border = BORDER_BOTTOM

section_header(ws_d, "Cash Flow Available for Debt Service", DR['beg_cash'] - 1, span=7)
label(ws_d, DR['beg_cash'], "Beginning cash balance")
val(ws_d, DR['beg_cash'], 2, f"='Sources & Uses'!B{SU['use_min_cash']}", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    prev_col = 'B' if i == 0 else PROJ_COLS[i - 1]
    src = f"{prev_col}{DR['beg_cash']}" if i == 0 else f"{prev_col}{DR['end_cash']}"
    val(ws_d, DR['beg_cash'], 3 + i, f"={src}", BLACK, FMT_USD_MM)

label(ws_d, DR['cfads'], "Levered FCF before debt service (CFADS)")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['cfads'], 3 + i, f"='Free Cash Flow'!{col}{FR['cfads']}", GREEN, FMT_USD_MM)

label(ws_d, DR['mand_amort_tlb'], "Less: mandatory amortization — Term Loan B")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['mand_amort_tlb'], 3 + i, f"={col}{DR['tlb_amort']}", BLACK, FMT_USD_MM)

label(ws_d, DR['mand_amort_notes'], "Less: mandatory amortization — Senior Notes (bullet)")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['mand_amort_notes'], 3 + i, 0, BLACK, FMT_USD_MM)

label(ws_d, DR['cf_before_revolver'], "Cash flow before revolver activity (vs. min. cash target)", bold=True)
for i, col in enumerate(PROJ_COLS):
    f = f"={col}{DR['beg_cash']}+{col}{DR['cfads']}-{col}{DR['mand_amort_tlb']}-{col}{DR['mand_amort_notes']}-Assumptions!$B${AR['min_cash']}/1000"
    val(ws_d, DR['cf_before_revolver'], 3 + i, f, BOLD_BLACK, FMT_USD_MM)

section_header(ws_d, "Revolver", DR['rev_beg'] - 1, span=7)
label(ws_d, DR['rev_beg'], "Beginning balance")
val(ws_d, DR['rev_beg'], 2, f"='Sources & Uses'!B{SU['src_revolver']}", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    prev_col = 'B' if i == 0 else PROJ_COLS[i - 1]
    src = f"{prev_col}{DR['rev_beg']}" if i == 0 else f"{prev_col}{DR['rev_end']}"
    val(ws_d, DR['rev_beg'], 3 + i, f"={src}", BLACK, FMT_USD_MM)

label(ws_d, DR['rev_draw'], "Draw / (Paydown)")
for i, col in enumerate(PROJ_COLS):
    f = f"=IF({col}{DR['cf_before_revolver']}<0,-{col}{DR['cf_before_revolver']},-MIN({col}{DR['rev_beg']},MAX(0,{col}{DR['cf_before_revolver']})))"
    val(ws_d, DR['rev_draw'], 3 + i, f, BLACK, FMT_USD_MM)

label(ws_d, DR['rev_end'], "Ending balance")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['rev_end'], 3 + i, f"={col}{DR['rev_beg']}+{col}{DR['rev_draw']}", BLACK, FMT_USD_MM)

label(ws_d, DR['rev_avg'], "Average balance (memo only; interest uses beginning balance below)")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['rev_avg'], 3 + i, f"=({col}{DR['rev_beg']}+{col}{DR['rev_end']})/2", BLACK, FMT_USD_MM)

label(ws_d, DR['rev_int'], "Interest expense on beginning balance (incl. commitment fee on undrawn)")
for i, col in enumerate(PROJ_COLS):
    f = f"={col}{DR['rev_beg']}*Assumptions!$B${AR['rev_rate']}+(Assumptions!$B${AR['revolver_cap']}/1000-{col}{DR['rev_beg']})*Assumptions!$B${AR['rev_commit_fee']}"
    val(ws_d, DR['rev_int'], 3 + i, f, BLACK, FMT_USD_MM)

section_header(ws_d, "Term Loan B", DR['tlb_beg'] - 1, span=7)
label(ws_d, DR['tlb_beg'], "Beginning balance")
val(ws_d, DR['tlb_beg'], 2, f"='Sources & Uses'!B{SU['src_tlb']}", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    prev_col = 'B' if i == 0 else PROJ_COLS[i - 1]
    src = f"{prev_col}{DR['tlb_beg']}" if i == 0 else f"{prev_col}{DR['tlb_end']}"
    val(ws_d, DR['tlb_beg'], 3 + i, f"={src}", BLACK, FMT_USD_MM)

label(ws_d, DR['tlb_amort'], "Mandatory amortization (1.00% of original principal p.a.)")
for i, col in enumerate(PROJ_COLS):
    f = f"=MIN({col}{DR['tlb_beg']},Assumptions!$B${AR['tlb_amort_pct']}*$B${DR['tlb_beg']})"
    val(ws_d, DR['tlb_amort'], 3 + i, f, BLACK, FMT_USD_MM)

label(ws_d, DR['tlb_sweep'], "Optional cash flow sweep")
for i, col in enumerate(PROJ_COLS):
    f = (f"=MIN(MAX(0,{col}{DR['tlb_beg']}-{col}{DR['tlb_amort']}),"
         f"Assumptions!$B${AR['sweep_pct']}*MAX(0,{col}{DR['cf_before_revolver']}-{col}{DR['rev_beg']}))")
    val(ws_d, DR['tlb_sweep'], 3 + i, f, BLACK, FMT_USD_MM)

label(ws_d, DR['tlb_end'], "Ending balance")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['tlb_end'], 3 + i, f"={col}{DR['tlb_beg']}-{col}{DR['tlb_amort']}-{col}{DR['tlb_sweep']}", BLACK, FMT_USD_MM)

label(ws_d, DR['tlb_avg'], "Average balance (memo only; interest uses beginning balance below)")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['tlb_avg'], 3 + i, f"=({col}{DR['tlb_beg']}+{col}{DR['tlb_end']})/2", BLACK, FMT_USD_MM)

label(ws_d, DR['tlb_int'], "Interest expense on beginning balance")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['tlb_int'], 3 + i, f"={col}{DR['tlb_beg']}*Assumptions!$B${AR['tlb_rate']}", BLACK, FMT_USD_MM)

section_header(ws_d, "Senior Notes", DR['notes_beg'] - 1, span=7)
label(ws_d, DR['notes_beg'], "Beginning balance")
val(ws_d, DR['notes_beg'], 2, f"='Sources & Uses'!B{SU['src_notes']}", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    prev_col = 'B' if i == 0 else PROJ_COLS[i - 1]
    src = f"{prev_col}{DR['notes_beg']}" if i == 0 else f"{prev_col}{DR['notes_end']}"
    val(ws_d, DR['notes_beg'], 3 + i, f"={src}", BLACK, FMT_USD_MM)

label(ws_d, DR['notes_amort'], "Mandatory amortization (bullet at maturity)")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['notes_amort'], 3 + i, 0, BLACK, FMT_USD_MM)

label(ws_d, DR['notes_end'], "Ending balance")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['notes_end'], 3 + i, f"={col}{DR['notes_beg']}-{col}{DR['notes_amort']}", BLACK, FMT_USD_MM)

label(ws_d, DR['notes_avg'], "Average balance (memo only; interest uses beginning balance below)")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['notes_avg'], 3 + i, f"=({col}{DR['notes_beg']}+{col}{DR['notes_end']})/2", BLACK, FMT_USD_MM)

label(ws_d, DR['notes_int'], "Interest expense on beginning balance")
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['notes_int'], 3 + i, f"={col}{DR['notes_beg']}*Assumptions!$B${AR['notes_rate']}", BLACK, FMT_USD_MM)

section_header(ws_d, "Consolidated Debt & Cash", DR['total_int'] - 1, span=7)
label(ws_d, DR['total_int'], "Total interest expense", bold=True)
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['total_int'], 3 + i, f"={col}{DR['rev_int']}+{col}{DR['tlb_int']}+{col}{DR['notes_int']}", BOLD_BLACK, FMT_USD_MM)

label(ws_d, DR['total_debt'], "Total debt outstanding, ending", bold=True)
val(ws_d, DR['total_debt'], 2, f"=B{DR['rev_beg']}+B{DR['tlb_beg']}+B{DR['notes_beg']}", BOLD_BLACK, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['total_debt'], 3 + i, f"={col}{DR['rev_end']}+{col}{DR['tlb_end']}+{col}{DR['notes_end']}", BOLD_BLACK, FMT_USD_MM)

label(ws_d, DR['end_cash'], "Ending cash balance")
val(ws_d, DR['end_cash'], 2, f"=B{DR['beg_cash']}", BLACK, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    f = f"={col}{DR['beg_cash']}+{col}{DR['cfads']}-{col}{DR['mand_amort_tlb']}-{col}{DR['mand_amort_notes']}+{col}{DR['rev_draw']}-{col}{DR['tlb_sweep']}"
    val(ws_d, DR['end_cash'], 3 + i, f, BLACK, FMT_USD_MM)

label(ws_d, DR['net_debt'], "Total net debt, ending", bold=True)
val(ws_d, DR['net_debt'], 2, f"=B{DR['total_debt']}-B{DR['end_cash']}", BOLD_BLACK, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_d, DR['net_debt'], 3 + i, f"={col}{DR['total_debt']}-{col}{DR['end_cash']}", BOLD_BLACK, FMT_USD_MM)

label(ws_d, DR['check_close'], "CHECK: net debt at close vs. Sources & Uses opening net debt (must = 0)", bold=True)
cf = f"=B{DR['net_debt']}-'Sources & Uses'!B{SU['opening_net_debt']}"
val(ws_d, DR['check_close'], 2, cf, BOLD_BLACK, FMT_USD_MM)
ws_d.conditional_formatting.add(f"B{DR['check_close']}", CellIsRule(operator='notEqual', formula=['0'], fill=CHECK_BAD_FILL))
ws_d.conditional_formatting.add(f"B{DR['check_close']}", CellIsRule(operator='equal', formula=['0'], fill=CHECK_OK_FILL))

print("Debt Schedule tab built.")

# =============================================================================
# 6. FREE CASH FLOW TAB
# =============================================================================
set_col_widths(ws_fcf, [46, 13, 13, 13, 13, 13, 13])
title_bar(ws_fcf, "V2X, Inc. (NYSE: VVX) — Free Cash Flow", span=7)
subtitle(ws_fcf, "$ in millions. Levered FCF available for debt paydown, feeding the Debt Schedule waterfall.", 2, span=7)

label(ws_fcf, FR['header'], "")
for i, y in enumerate(PROJ_YEARS):
    val(ws_fcf, FR['header'], 3 + i, f"FY{y}", BOLD_BLACK)
    ws_fcf.cell(row=FR['header'], column=3 + i).alignment = Alignment(horizontal="center")
    ws_fcf.cell(row=FR['header'], column=3 + i).border = BORDER_BOTTOM

section_header(ws_fcf, "Levered Free Cash Flow", FR['ni'] - 1, span=7)
label(ws_fcf, FR['ni'], "Net income")
for i, col in enumerate(PROJ_COLS):
    val(ws_fcf, FR['ni'], 3 + i, f"='Operating Model'!{col}{OM['ni']}", GREEN, FMT_USD_MM)

label(ws_fcf, FR['da'], "+ D&A")
for i, col in enumerate(PROJ_COLS):
    val(ws_fcf, FR['da'], 3 + i, f"='Operating Model'!{col}{OM['da']}", GREEN, FMT_USD_MM)

label(ws_fcf, FR['capex'], "− Capex")
for i, col in enumerate(PROJ_COLS):
    val(ws_fcf, FR['capex'], 3 + i, f"='Operating Model'!{col}{OM['capex']}", GREEN, FMT_USD_MM)

label(ws_fcf, FR['nwc_cash'], "+/− Cash impact from NWC")
for i, col in enumerate(PROJ_COLS):
    val(ws_fcf, FR['nwc_cash'], 3 + i, f"='Operating Model'!{col}{OM['nwc_cash']}", GREEN, FMT_USD_MM)

label(ws_fcf, FR['cfads'], "Levered FCF before debt service (CFADS)", bold=True)
for i, col in enumerate(PROJ_COLS):
    f = f"={col}{FR['ni']}+{col}{FR['da']}-{col}{FR['capex']}+{col}{FR['nwc_cash']}"
    val(ws_fcf, FR['cfads'], 3 + i, f, BOLD_BLACK, FMT_USD_MM)

section_header(ws_fcf, "Debt Service (from Debt Schedule)", FR['mand_amort'] - 1, span=7)
label(ws_fcf, FR['mand_amort'], "Less: mandatory amortization (TLB + Notes)")
for i, col in enumerate(PROJ_COLS):
    f = f"='Debt Schedule'!{col}{DR['mand_amort_tlb']}+'Debt Schedule'!{col}{DR['mand_amort_notes']}"
    val(ws_fcf, FR['mand_amort'], 3 + i, f, GREEN, FMT_USD_MM)

label(ws_fcf, FR['sweep'], "Less: optional cash sweep (Term Loan B)")
for i, col in enumerate(PROJ_COLS):
    val(ws_fcf, FR['sweep'], 3 + i, f"='Debt Schedule'!{col}{DR['tlb_sweep']}", GREEN, FMT_USD_MM)

label(ws_fcf, FR['revolver'], "+/− Net revolver draw / (paydown)")
for i, col in enumerate(PROJ_COLS):
    val(ws_fcf, FR['revolver'], 3 + i, f"='Debt Schedule'!{col}{DR['rev_draw']}", GREEN, FMT_USD_MM)

label(ws_fcf, FR['net_chg_cash'], "Net change in cash", bold=True)
for i, col in enumerate(PROJ_COLS):
    f = f"={col}{FR['cfads']}-{col}{FR['mand_amort']}-{col}{FR['sweep']}+{col}{FR['revolver']}"
    val(ws_fcf, FR['net_chg_cash'], 3 + i, f, BOLD_BLACK, FMT_USD_MM)

label(ws_fcf, FR['check'], "CHECK: net change in cash vs. Debt Schedule roll-forward (must = 0)", bold=True)
for i, col in enumerate(PROJ_COLS):
    f = f"={col}{FR['net_chg_cash']}-('Debt Schedule'!{col}{DR['end_cash']}-'Debt Schedule'!{col}{DR['beg_cash']})"
    val(ws_fcf, FR['check'], 3 + i, f, BOLD_BLACK, FMT_USD_MM)
    cl = col
    ws_fcf.conditional_formatting.add(f"{cl}{FR['check']}", CellIsRule(operator='notEqual', formula=['0'], fill=CHECK_BAD_FILL))
    ws_fcf.conditional_formatting.add(f"{cl}{FR['check']}", CellIsRule(operator='equal', formula=['0'], fill=CHECK_OK_FILL))

print("Free Cash Flow tab built.")

# =============================================================================
# 7. RETURNS TAB
# =============================================================================
set_col_widths(ws_r, [50, 14, 14, 14, 14, 14, 14, 14])
title_bar(ws_r, "V2X, Inc. (NYSE: VVX) — Returns", span=8)
subtitle(ws_r, "$ in millions except per-x and % rows.", 2, span=8)

RR = {}
r = 4
section_header(ws_r, "Entry (12/31/2025)", r, span=8); r += 1
RR['entry_ebitda'] = r; label(ws_r, r, "Entry EBITDA")
val(ws_r, r, 2, f"='Sources & Uses'!B{SU['entry_ebitda']}", GREEN, FMT_USD_MM); r += 1
RR['entry_mult'] = r; label(ws_r, r, "Entry multiple")
val(ws_r, r, 2, f"='Sources & Uses'!B{SU['entry_mult']}", GREEN, FMT_X); r += 1
RR['entry_tev'] = r; label(ws_r, r, "Entry TEV")
val(ws_r, r, 2, f"='Sources & Uses'!B{SU['entry_tev']}", GREEN, FMT_USD_MM); r += 1
RR['sponsor_equity'] = r; label(ws_r, r, "Sponsor equity investment", bold=True)
val(ws_r, r, 2, f"='Sources & Uses'!B{SU['src_equity']}", GREEN, FMT_USD_MM, bold=True); r += 2

section_header(ws_r, f"Exit (FY{PROJ_YEARS[-1]})", r, span=8); r += 1
RR['exit_ebitda'] = r; label(ws_r, r, "Exit EBITDA")
val(ws_r, r, 2, f"='Operating Model'!G{OM['ebitda']}", GREEN, FMT_USD_MM); r += 1
RR['exit_mult'] = r; label(ws_r, r, "Exit multiple")
val(ws_r, r, 2, f"=Assumptions!B{AR['exit_mult']}", GREEN, FMT_X); r += 1
RR['exit_tev'] = r; label(ws_r, r, "Exit TEV")
val(ws_r, r, 2, f"=B{RR['exit_ebitda']}*B{RR['exit_mult']}", BLACK, FMT_USD_MM); r += 1
RR['exit_net_debt'] = r; label(ws_r, r, "Exit net debt")
val(ws_r, r, 2, f"='Debt Schedule'!G{DR['net_debt']}", GREEN, FMT_USD_MM); r += 1
RR['exit_equity'] = r; label(ws_r, r, "Exit equity value", bold=True)
val(ws_r, r, 2, f"=B{RR['exit_tev']}-B{RR['exit_net_debt']}", BOLD_BLACK, FMT_USD_MM); r += 2

section_header(ws_r, "Sponsor Returns", r, span=8); r += 1
RR['moic'] = r; label(ws_r, r, "MOIC", bold=True)
val(ws_r, r, 2, f"=B{RR['exit_equity']}/B{RR['sponsor_equity']}", BOLD_BLACK, FMT_X); r += 1
RR['irr'] = r; label(ws_r, r, "IRR", bold=True)
r += 2

section_header(ws_r, "Cash Flow Timeline for IRR", r, span=8); r += 1
RR['cf_years'] = r
label(ws_r, r, "Year")
for i in range(HOLD_YEARS + 1):
    val(ws_r, r, 2 + i, i, BOLD_BLACK, FMT_YEAR)
    ws_r.cell(row=r, column=2 + i).alignment = Alignment(horizontal="center")
r += 1
RR['cf_row'] = r
label(ws_r, r, "Sponsor cash flow", bold=True)
val(ws_r, r, 2, f"=-B{RR['sponsor_equity']}", BOLD_BLACK, FMT_USD_MM)
for i in range(1, HOLD_YEARS):
    val(ws_r, r, 2 + i, 0, BOLD_BLACK, FMT_USD_MM)
val(ws_r, r, 2 + HOLD_YEARS, f"=B{RR['exit_equity']}", BOLD_BLACK, FMT_USD_MM)
r += 1

# now fix the IRR formula written earlier with the correct row
last_cf_col = get_column_letter(2 + HOLD_YEARS)
ws_r.cell(row=RR['irr'], column=2, value=f"=IRR(B{RR['cf_row']}:{last_cf_col}{RR['cf_row']})")
ws_r.cell(row=RR['irr'], column=2).number_format = FMT_PCT
ws_r.cell(row=RR['irr'], column=2).font = BOLD_BLACK
r += 1

section_header(ws_r, "Returns Attribution — Value Creation Bridge (ties exactly to Exit Equity Value)", r, span=8); r += 1
RR['attr_start'] = r
label(ws_r, r, "Sponsor equity (entry)")
val(ws_r, r, 2, f"=B{RR['sponsor_equity']}", GREEN, FMT_USD_MM); r += 1
RR['attr_ebitda_growth'] = r
label(ws_r, r, "+ EBITDA growth (Entry multiple x change in EBITDA)")
val(ws_r, r, 2, f"=B{RR['entry_mult']}*(B{RR['exit_ebitda']}-B{RR['entry_ebitda']})", BLACK, FMT_USD_MM); r += 1
RR['attr_multiple'] = r
label(ws_r, r, "+ Multiple expansion / (contraction)")
val(ws_r, r, 2, f"=(B{RR['exit_mult']}-B{RR['entry_mult']})*B{RR['exit_ebitda']}", BLACK, FMT_USD_MM); r += 1
RR['attr_delever'] = r
label(ws_r, r, "+ Deleveraging (opening net debt less exit net debt)")
val(ws_r, r, 2, f"='Sources & Uses'!B{SU['opening_net_debt']}-B{RR['exit_net_debt']}", GREEN, FMT_USD_MM); r += 1
RR['attr_fees'] = r
label(ws_r, r, "− Fees & frictional costs at entry")
val(ws_r, r, 2, f"=-('Sources & Uses'!B{SU['use_txn_fees']}+'Sources & Uses'!B{SU['use_fin_fees']})", GREEN, FMT_USD_MM); r += 1
RR['attr_total'] = r
label(ws_r, r, "Total: Exit equity value (attribution check)", bold=True)
val(ws_r, r, 2, f"=SUM(B{RR['attr_start']}:B{RR['attr_fees']})", BOLD_BLACK, FMT_USD_MM)
for col in range(1, 3):
    ws_r.cell(row=r, column=col).border = BORDER_TOP
r += 1
RR['attr_check'] = r
label(ws_r, r, "CHECK: attribution total − exit equity value (must = 0)", bold=True)
val(ws_r, r, 2, f"=B{RR['attr_total']}-B{RR['exit_equity']}", BOLD_BLACK, FMT_USD_MM)
ws_r.conditional_formatting.add(f"B{RR['attr_check']}", CellIsRule(operator='notEqual', formula=['0.01'], fill=CHECK_BAD_FILL))
ws_r.conditional_formatting.add(f"B{RR['attr_check']}", CellIsRule(operator='between', formula=['-0.01', '0.01'], fill=CHECK_OK_FILL))
r += 1

print("Returns tab built.")

# =============================================================================
# 8. CREDIT STATS TAB
# =============================================================================
set_col_widths(ws_cs, [40, 13, 13, 13, 13, 13, 13])
title_bar(ws_cs, "V2X, Inc. (NYSE: VVX) — Credit Stats", span=7)
subtitle(ws_cs, "$ in millions. Shows whether the LBO capital structure is financeable across the hold.", 2, span=7)

CSR = {}
r = 4
label(ws_cs, r, "")
val(ws_cs, r, 2, "At Close", BOLD_BLACK); ws_cs.cell(row=r, column=2).alignment = Alignment(horizontal="center")
for i, y in enumerate(PROJ_YEARS):
    val(ws_cs, r, 3 + i, f"FY{y}", BOLD_BLACK)
    ws_cs.cell(row=r, column=3 + i).alignment = Alignment(horizontal="center")
    ws_cs.cell(row=r, column=3 + i).border = BORDER_BOTTOM
ws_cs.cell(row=r, column=2).border = BORDER_BOTTOM
r += 1

CSR['total_debt'] = r; label(ws_cs, r, "Total debt outstanding")
val(ws_cs, r, 2, f"='Debt Schedule'!B{DR['total_debt']}", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_cs, r, 3 + i, f"='Debt Schedule'!{col}{DR['total_debt']}", GREEN, FMT_USD_MM)
r += 1

CSR['net_debt'] = r; label(ws_cs, r, "Net debt")
val(ws_cs, r, 2, f"='Debt Schedule'!B{DR['net_debt']}", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_cs, r, 3 + i, f"='Debt Schedule'!{col}{DR['net_debt']}", GREEN, FMT_USD_MM)
r += 1

CSR['ebitda'] = r; label(ws_cs, r, "EBITDA")
val(ws_cs, r, 2, f"='Sources & Uses'!B{SU['entry_ebitda']}", GREEN, FMT_USD_MM)
for i, col in enumerate(PROJ_COLS):
    val(ws_cs, r, 3 + i, f"='Operating Model'!{col}{OM['ebitda']}", GREEN, FMT_USD_MM)
r += 1

CSR['interest'] = r; label(ws_cs, r, "Total interest expense")
val(ws_cs, r, 2, "n/a", ITALIC_GRAY)
for i, col in enumerate(PROJ_COLS):
    val(ws_cs, r, 3 + i, f"='Debt Schedule'!{col}{DR['total_int']}", GREEN, FMT_USD_MM)
r += 2

CSR['debt_ebitda'] = r; label(ws_cs, r, "Total Debt / EBITDA", bold=True)
val(ws_cs, r, 2, f"=B{CSR['total_debt']}/B{CSR['ebitda']}", BOLD_BLACK, FMT_X)
for i, col in enumerate(PROJ_COLS):
    val(ws_cs, r, 3 + i, f"={col}{CSR['total_debt']}/{col}{CSR['ebitda']}", BOLD_BLACK, FMT_X)
r += 1

CSR['netdebt_ebitda'] = r; label(ws_cs, r, "Net Debt / EBITDA", bold=True)
val(ws_cs, r, 2, f"=B{CSR['net_debt']}/B{CSR['ebitda']}", BOLD_BLACK, FMT_X)
for i, col in enumerate(PROJ_COLS):
    val(ws_cs, r, 3 + i, f"={col}{CSR['net_debt']}/{col}{CSR['ebitda']}", BOLD_BLACK, FMT_X)
r += 1

CSR['ebitda_interest'] = r; label(ws_cs, r, "EBITDA / Interest expense", bold=True)
val(ws_cs, r, 2, "n/a", ITALIC_GRAY)
for i, col in enumerate(PROJ_COLS):
    val(ws_cs, r, 3 + i, f"={col}{CSR['ebitda']}/{col}{CSR['interest']}", BOLD_BLACK, FMT_X)
r += 1

print("Credit Stats tab built.")

# =============================================================================
# 9. SENSITIVITIES TAB
# =============================================================================
set_col_widths(ws_s, [26, 13, 13, 13, 13, 13, 13, 13])
title_bar(ws_s, "V2X, Inc. (NYSE: VVX) — Sensitivities", span=8)
subtitle(ws_s, "Every cell is a live formula. Change any header (blue) to re-run the grid.", 2, span=8)

ENTRY_GRID = [9.0, 10.0, ENTRY_MULTIPLE, 12.0, 13.0]
EXIT_GRID = [9.0, 10.0, EXIT_MULTIPLE, 12.0, 13.0]
CAGR_GRID = [0.020, 0.030, 0.043, 0.055, 0.070]

r = 4
section_header(ws_s, "Table 1 — IRR by Entry Multiple (rows) x Exit Multiple (columns)", r, span=8); r += 1
t1_hdr_row = r
label(ws_s, r, "Entry \\ Exit")
for j, xm in enumerate(EXIT_GRID):
    val(ws_s, r, 3 + j, xm, BLUE, FMT_X)
r += 1
t1_first_data_row = r
for i, em in enumerate(ENTRY_GRID):
    row = t1_first_data_row + i
    val(ws_s, row, 2, em, BLUE, FMT_X)
    for j in range(len(EXIT_GRID)):
        col = get_column_letter(3 + j)
        hdr_col = get_column_letter(3 + j)
        f = (f"=(({hdr_col}${t1_hdr_row}*'Operating Model'!$G${OM['ebitda']}-'Debt Schedule'!$G${DR['net_debt']})"
             f"/($B{row}*'Sources & Uses'!$B${SU['entry_ebitda']}*(1+Assumptions!$B${AR['txn_fee_pct']})"
             f"+'Sources & Uses'!$B${SU['use_fin_fees']}+'Sources & Uses'!$B${SU['use_min_cash']}"
             f"-'Sources & Uses'!$B${SU['new_tlb']}-'Sources & Uses'!$B${SU['new_notes']}))"
             f"^(1/Assumptions!$B${AR['hold_years']})-1")
        c = val(ws_s, row, 3 + j, f, BLACK, FMT_PCT)
        if abs(em - ENTRY_MULTIPLE) < 1e-9 and abs(EXIT_GRID[j] - EXIT_MULTIPLE) < 1e-9:
            c.fill = SECTION_FILL
r += len(ENTRY_GRID) + 2

section_header(ws_s, "Table 2 — IRR by Exit Multiple (rows) x 5-Yr Revenue CAGR (columns)", r, span=8); r += 1
ws_s.cell(row=r, column=1, value="Simplification: aggregate (non-tranche) debt paydown; interest expense held at base-case $ (2nd-order effect of paydown speed on interest is not re-derived per scenario); same margin/capex/NWC/tax assumptions as base case. See helper block below.").font = ITALIC_GRAY
ws_s.row_dimensions[r].height = 28
r += 1
t2_hdr_row = r
label(ws_s, r, "Exit Mult \\ CAGR")
for j, cg in enumerate(CAGR_GRID):
    val(ws_s, r, 3 + j, cg, BLUE, FMT_PCT)
r += 1
t2_first_data_row = r
for i, xm in enumerate(EXIT_GRID):
    row = t2_first_data_row + i
    val(ws_s, row, 2, xm, BLUE, FMT_X)
r += len(EXIT_GRID) + 2

# --- Helper block: simplified aggregate debt roll-forward per CAGR scenario ---
section_header(ws_s, "Sensitivity Helper — Simplified 5-Year Roll-Forward by Revenue CAGR Scenario (feeds Table 2)", r, span=8); r += 1
helper_start = r
scenario_exit_ebitda_row = {}
scenario_exit_netdebt_row = {}

for s_idx, cagr_col in enumerate(['C', 'D', 'E', 'F', 'G']):
    label(ws_s, r, f"Scenario {s_idx + 1}: Revenue CAGR = {cagr_col}{t2_hdr_row}", bold=True)
    r += 1
    yr_label_row = r
    for i, y in enumerate(PROJ_YEARS):
        val(ws_s, r, 3 + i, f"FY{y}", BOLD_BLACK)
    r += 1
    rev_row = r; label(ws_s, r, "Revenue")
    val(ws_s, r, 2, f"='Historical Financials'!{HR_COL_2025_HIST}{HR['revenue']}/1000", GREEN, FMT_USD_MM)
    for i, col in enumerate(PROJ_COLS):
        prev = 'B' if i == 0 else PROJ_COLS[i - 1]
        val(ws_s, r, 3 + i, f"={prev}{r}*(1+${cagr_col}${t2_hdr_row})", BLACK, FMT_USD_MM)
    r += 1
    ebitda_row = r; label(ws_s, r, "EBITDA")
    for i, col in enumerate(PROJ_COLS):
        val(ws_s, r, 3 + i, f"={col}{rev_row}*Assumptions!{ASSUMP_COLS[i]}{AR['ebitda_margin_row']}", BLACK, FMT_USD_MM)
    r += 1
    da_row = r; label(ws_s, r, "D&A")
    for i, col in enumerate(PROJ_COLS):
        val(ws_s, r, 3 + i, f"={col}{rev_row}*Assumptions!$B${AR['da_pct']}", BLACK, FMT_USD_MM)
    r += 1
    ebt_row = r; label(ws_s, r, "EBT (interest held at base-case $)")
    for i, col in enumerate(PROJ_COLS):
        val(ws_s, r, 3 + i, f"={col}{ebitda_row}-{col}{da_row}-'Debt Schedule'!{col}{DR['total_int']}", BLACK, FMT_USD_MM)
    r += 1
    ni_row = r; label(ws_s, r, "Net income")
    for i, col in enumerate(PROJ_COLS):
        val(ws_s, r, 3 + i, f"={col}{ebt_row}-MAX(0,{col}{ebt_row})*Assumptions!$B${AR['tax_rate']}", BLACK, FMT_USD_MM)
    r += 1
    fcf_row = r; label(ws_s, r, "FCF before debt service")
    for i, col in enumerate(PROJ_COLS):
        prev = 'B' if i == 0 else PROJ_COLS[i - 1]
        f = (f"={col}{ni_row}+{col}{da_row}-{col}{rev_row}*Assumptions!$B${AR['capex_pct']}"
             f"-({col}{rev_row}-{prev}{rev_row})*Assumptions!$B${AR['nwc_pct']}")
        val(ws_s, r, 3 + i, f, BLACK, FMT_USD_MM)
    r += 1
    debtbeg_row = r; label(ws_s, r, "Debt, beginning")
    val(ws_s, r, 2, f"='Sources & Uses'!B{SU['new_debt_total']}", GREEN, FMT_USD_MM)
    for i, col in enumerate(PROJ_COLS):
        prev = 'B' if i == 0 else PROJ_COLS[i - 1]
        src = f"{prev}{debtbeg_row}" if i == 0 else f"{prev}{debtbeg_row + 2}"
        val(ws_s, r, 3 + i, f"={src}", BLACK, FMT_USD_MM)
    r += 1
    sweep_row = r; label(ws_s, r, "Mandatory amort. + sweep")
    for i, col in enumerate(PROJ_COLS):
        f = (f"=MIN(MAX(0,{col}{debtbeg_row}-'Debt Schedule'!{col}{DR['tlb_amort']}),"
             f"'Debt Schedule'!{col}{DR['tlb_amort']}+Assumptions!$B${AR['sweep_pct']}*MAX(0,{col}{fcf_row}-'Debt Schedule'!{col}{DR['tlb_amort']}))")
        val(ws_s, r, 3 + i, f, BLACK, FMT_USD_MM)
    r += 1
    debtend_row = r; label(ws_s, r, "Debt, ending")
    for i, col in enumerate(PROJ_COLS):
        val(ws_s, r, 3 + i, f"={col}{debtbeg_row}-{col}{sweep_row}", BLACK, FMT_USD_MM)
    r += 1
    cashbeg_row = r; label(ws_s, r, "Cash, beginning")
    val(ws_s, r, 2, f"=Assumptions!$B${AR['min_cash']}/1000", GREEN, FMT_USD_MM)
    for i, col in enumerate(PROJ_COLS):
        prev = 'B' if i == 0 else PROJ_COLS[i - 1]
        src = f"{prev}{cashbeg_row}" if i == 0 else f"{prev}{cashbeg_row + 1}"
        val(ws_s, r, 3 + i, f"={src}", BLACK, FMT_USD_MM)
    r += 1
    cashend_row = r; label(ws_s, r, "Cash, ending")
    for i, col in enumerate(PROJ_COLS):
        val(ws_s, r, 3 + i, f"={col}{cashbeg_row}+{col}{fcf_row}-{col}{sweep_row}", BLACK, FMT_USD_MM)
    r += 2

    scenario_exit_ebitda_row[cagr_col] = ebitda_row
    scenario_exit_netdebt_row[cagr_col] = (debtend_row, cashend_row)

# Now populate Table 2 interior using the helper block's Year-5 (column G) results
for i, xm in enumerate(EXIT_GRID):
    row = t2_first_data_row + i
    for j, cagr_col in enumerate(['C', 'D', 'E', 'F', 'G']):
        ebitda_r = scenario_exit_ebitda_row[cagr_col]
        debtend_r, cashend_r = scenario_exit_netdebt_row[cagr_col]
        col_letter = get_column_letter(3 + j)
        f = (f"=(($B{row}*G{ebitda_r}-(G{debtend_r}-G{cashend_r}))"
             f"/'Sources & Uses'!$B${SU['src_equity']})^(1/Assumptions!$B${AR['hold_years']})-1")
        c = val(ws_s, row, 3 + j, f, BLACK, FMT_PCT)
        if abs(xm - EXIT_MULTIPLE) < 1e-9 and j == 2:
            c.fill = SECTION_FILL

print("Sensitivities tab built.")

import os
for ws in wb.worksheets:
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

os.makedirs("output", exist_ok=True)
wb.save("output/V2X_LBO.xlsx")
print("Workbook saved to output/V2X_LBO.xlsx")

# =============================================================================
# 10. INDEPENDENT PYTHON NUMERIC REPLICA (verification + terminal printout)
#     Mirrors the exact same formulas written into the workbook above, using
#     the same Python constants, with its own fixed-point iteration to
#     resolve the interest <-> FCF <-> debt-paydown circularity (equivalent
#     to Excel's iterative calculation engine).
# =============================================================================

def compute_model():
    entry_ebitda = ENTRY_EBITDA / 1000.0  # $mm
    entry_tev = ENTRY_MULTIPLE * entry_ebitda
    existing_debt = EXISTING_DEBT_PRINCIPAL / 1000.0
    existing_cash = EXISTING_CASH / 1000.0
    existing_net_debt = existing_debt - existing_cash
    equity_purchase_price = entry_tev - existing_net_debt
    new_tlb = entry_ebitda * TLB_TURNS
    new_notes = entry_ebitda * NOTES_TURNS
    new_debt_total = new_tlb + new_notes
    txn_fees = entry_tev * TXN_FEE_PCT
    fin_fees = new_debt_total * FINANCING_FEE_PCT
    min_cash = MIN_CASH / 1000.0
    total_uses = equity_purchase_price + existing_debt + txn_fees + fin_fees + min_cash
    total_sources_excl_equity = existing_cash + new_tlb + new_notes + 0.0
    sponsor_equity = total_uses - total_sources_excl_equity
    opening_net_debt = new_debt_total - min_cash

    revenue, ebitda, da, capex, nwc = {}, {}, {}, {}, {}
    rev_prev = REVENUE[2025] / 1000.0
    nwc_prev0 = (RECEIVABLES[2025] + PREPAID_OTHER_CA[2025] - AP[2025] - COMP_BENEFITS[2025] - OTHER_ACCRUED[2025]) / 1000.0
    for idx, y in enumerate(PROJ_YEARS):
        rev = rev_prev * (1 + REV_GROWTH[idx])
        revenue[y] = rev
        ebitda[y] = rev * EBITDA_MARGIN[idx]
        da[y] = rev * DA_PCT_REV
        capex[y] = rev * CAPEX_PCT_REV
        nwc[y] = rev * NWC_PCT_REV
        rev_prev = rev

    total_int = {y: 0.0 for y in PROJ_YEARS}
    tlb_beg, tlb_end = {}, {}
    notes_beg, notes_end = {}, {}
    rev_beg, rev_end = {}, {}
    cash_beg, cash_end = {}, {}
    ni_d, fcf_d, tax_d, ebt_d = {}, {}, {}, {}

    for _ in range(300):
        prev_total_int = dict(total_int)
        tlb_prev, notes_prev, revolver_prev, cash_prev = new_tlb, new_notes, 0.0, min_cash
        nwc_prev_y = nwc_prev0
        for y in PROJ_YEARS:
            ebt = ebitda[y] - da[y] - total_int[y]
            tax = max(0, ebt) * TAX_RATE
            ni = ebt - tax
            nwc_cash = -(nwc[y] - nwc_prev_y)
            fcf = ni + da[y] - capex[y] + nwc_cash
            mand_amort_tlb = min(tlb_prev, TLB_MAND_AMORT_PCT * new_tlb)
            cf_before_revolver = cash_prev + fcf - mand_amort_tlb - min_cash
            if cf_before_revolver < 0:
                draw_paydown = -cf_before_revolver
            else:
                draw_paydown = -min(revolver_prev, max(0, cf_before_revolver))
            revolver_end = revolver_prev + draw_paydown
            cash_avail_sweep = max(0, cf_before_revolver - revolver_prev)
            sweep = min(max(0, tlb_prev - mand_amort_tlb), CASH_SWEEP_PCT * cash_avail_sweep)
            tlb_end_y = tlb_prev - mand_amort_tlb - sweep
            notes_end_y = notes_prev
            cash_end_y = cash_prev + fcf - mand_amort_tlb + draw_paydown - sweep

            tlb_beg[y], tlb_end[y] = tlb_prev, tlb_end_y
            notes_beg[y], notes_end[y] = notes_prev, notes_end_y
            rev_beg[y], rev_end[y] = revolver_prev, revolver_end
            cash_beg[y], cash_end[y] = cash_prev, cash_end_y
            ni_d[y], fcf_d[y], tax_d[y], ebt_d[y] = ni, fcf, tax, ebt

            # Interest computed on BEGINNING-of-year balances (documented convention;
            # avoids circularity entirely -- matches the Excel formulas exactly).
            rev_int = revolver_prev * REVOLVER_RATE + (REVOLVER_CAPACITY / 1000.0 - revolver_prev) * REVOLVER_COMMIT_FEE
            tlb_int = tlb_prev * TLB_RATE
            notes_int = notes_prev * NOTES_RATE
            total_int[y] = rev_int + tlb_int + notes_int

            tlb_prev, notes_prev, revolver_prev, cash_prev = tlb_end_y, notes_end_y, revolver_end, cash_end_y
            nwc_prev_y = nwc[y]

        if max(abs(total_int[y] - prev_total_int[y]) for y in PROJ_YEARS) < 1e-10:
            break

    exit_year = PROJ_YEARS[-1]
    exit_ebitda = ebitda[exit_year]
    exit_tev = EXIT_MULTIPLE * exit_ebitda
    exit_total_debt = tlb_end[exit_year] + notes_end[exit_year] + rev_end[exit_year]
    exit_cash = cash_end[exit_year]
    exit_net_debt = exit_total_debt - exit_cash
    exit_equity = exit_tev - exit_net_debt
    moic = exit_equity / sponsor_equity
    irr = moic ** (1.0 / HOLD_YEARS) - 1.0

    attrib_ebitda_growth = ENTRY_MULTIPLE * (exit_ebitda - entry_ebitda)
    attrib_multiple = (EXIT_MULTIPLE - ENTRY_MULTIPLE) * exit_ebitda
    attrib_delever = opening_net_debt - exit_net_debt
    attrib_fees = -(txn_fees + fin_fees)
    attrib_total = sponsor_equity + attrib_ebitda_growth + attrib_multiple + attrib_delever + attrib_fees

    credit_stats = {}
    for y in PROJ_YEARS:
        total_debt_y = tlb_end[y] + notes_end[y] + rev_end[y]
        net_debt_y = total_debt_y - cash_end[y]
        credit_stats[y] = dict(
            total_debt=total_debt_y, net_debt=net_debt_y, ebitda=ebitda[y],
            debt_ebitda=total_debt_y / ebitda[y], netdebt_ebitda=net_debt_y / ebitda[y],
            interest=total_int[y], coverage=ebitda[y] / total_int[y],
        )

    return dict(
        entry_ebitda=entry_ebitda, entry_tev=entry_tev, existing_net_debt=existing_net_debt,
        equity_purchase_price=equity_purchase_price, new_tlb=new_tlb, new_notes=new_notes,
        new_debt_total=new_debt_total, txn_fees=txn_fees, fin_fees=fin_fees, min_cash=min_cash,
        total_uses=total_uses, sponsor_equity=sponsor_equity, opening_net_debt=opening_net_debt,
        revenue=revenue, ebitda=ebitda, exit_ebitda=exit_ebitda, exit_tev=exit_tev,
        exit_net_debt=exit_net_debt, exit_equity=exit_equity, moic=moic, irr=irr,
        attrib_ebitda_growth=attrib_ebitda_growth, attrib_multiple=attrib_multiple,
        attrib_delever=attrib_delever, attrib_fees=attrib_fees, attrib_total=attrib_total,
        credit_stats=credit_stats,
    )


def print_summary(m):
    print("\n" + "=" * 78)
    print("V2X, INC. (NYSE: VVX) — SPONSOR LBO MODEL: RETURNS SUMMARY")
    print("=" * 78)
    print(f"{'Entry EBITDA (FY2025A)':<45}{'$' + format(m['entry_ebitda'], ',.1f') + 'mm':>25}")
    print(f"{'Entry TEV/EBITDA multiple':<45}{str(ENTRY_MULTIPLE) + 'x':>25}")
    print(f"{'Entry TEV':<45}{'$' + format(m['entry_tev'], ',.1f') + 'mm':>25}")
    print(f"{'Sponsor equity investment':<45}{'$' + format(m['sponsor_equity'], ',.1f') + 'mm':>25}")
    print(f"{'Total new debt at close (' + str(TLB_TURNS + NOTES_TURNS) + 'x EBITDA)':<45}{'$' + format(m['new_debt_total'], ',.1f') + 'mm':>25}")
    print("-" * 78)
    print(f"{'Exit year':<45}{('FY' + str(PROJ_YEARS[-1])):>25}")
    print(f"{'Exit EBITDA':<45}{'$' + format(m['exit_ebitda'], ',.1f') + 'mm':>25}")
    print(f"{'Exit TEV/EBITDA multiple':<45}{str(EXIT_MULTIPLE) + 'x':>25}")
    print(f"{'Exit TEV':<45}{'$' + format(m['exit_tev'], ',.1f') + 'mm':>25}")
    print(f"{'Exit net debt':<45}{'$' + format(m['exit_net_debt'], ',.1f') + 'mm':>25}")
    print(f"{'Exit equity value':<45}{'$' + format(m['exit_equity'], ',.1f') + 'mm':>25}")
    print("-" * 78)
    print(f"{'MOIC':<45}{format(m['moic'], '.2f') + 'x':>25}")
    print(f"{'IRR':<45}{format(m['irr'] * 100, '.1f') + '%':>25}")
    print("-" * 78)
    print("Returns attribution (ties exactly to exit equity value):")
    print(f"  {'Sponsor equity (entry)':<43}{'$' + format(m['sponsor_equity'], ',.1f') + 'mm':>23}")
    print(f"  {'+ EBITDA growth':<43}{'$' + format(m['attrib_ebitda_growth'], ',.1f') + 'mm':>23}")
    print(f"  {'+ Multiple expansion/(contraction)':<43}{'$' + format(m['attrib_multiple'], ',.1f') + 'mm':>23}")
    print(f"  {'+ Deleveraging':<43}{'$' + format(m['attrib_delever'], ',.1f') + 'mm':>23}")
    print(f"  {'- Fees & frictional costs':<43}{'$' + format(m['attrib_fees'], ',.1f') + 'mm':>23}")
    print(f"  {'= Exit equity value (check)':<43}{'$' + format(m['attrib_total'], ',.1f') + 'mm':>23}")
    tie = abs(m['attrib_total'] - m['exit_equity']) < 0.05
    print(f"  {'Ties to exit equity value above:':<43}{'YES' if tie else 'NO -- CHECK MODEL':>23}")

    print("\n" + "=" * 78)
    print("CREDIT STATS")
    print("=" * 78)
    hdr = f"{'':<28}" + "".join(f"{'FY' + str(y):>12}" for y in PROJ_YEARS)
    print(hdr)
    cs = m['credit_stats']
    print(f"{'Total Debt ($mm)':<28}" + "".join(f"{cs[y]['total_debt']:>12,.0f}" for y in PROJ_YEARS))
    print(f"{'Net Debt ($mm)':<28}" + "".join(f"{cs[y]['net_debt']:>12,.0f}" for y in PROJ_YEARS))
    print(f"{'EBITDA ($mm)':<28}" + "".join(f"{cs[y]['ebitda']:>12,.1f}" for y in PROJ_YEARS))
    print(f"{'Total Debt / EBITDA':<28}" + "".join(f"{cs[y]['debt_ebitda']:>11.2f}x" for y in PROJ_YEARS))
    print(f"{'Net Debt / EBITDA':<28}" + "".join(f"{cs[y]['netdebt_ebitda']:>11.2f}x" for y in PROJ_YEARS))
    print(f"{'EBITDA / Interest':<28}" + "".join(f"{cs[y]['coverage']:>11.2f}x" for y in PROJ_YEARS))
    print("=" * 78 + "\n")


_model = compute_model()
print_summary(_model)
